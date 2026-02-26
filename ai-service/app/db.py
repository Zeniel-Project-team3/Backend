import json
import hashlib
from contextlib import contextmanager
from datetime import date, datetime
from typing import Any

import psycopg
from openpyxl import load_workbook

from app.config import settings


def _dsn() -> str:
    if settings.db_url:
        raw = settings.db_url.strip()
        if raw.startswith("jdbc:"):
            raw = raw[len("jdbc:") :]
        if "://" not in raw:
            return f"postgresql://{settings.db_user}:{settings.db_password}@{raw}"
        return raw
    return (
        f"host={settings.db_host} "
        f"port={settings.db_port} "
        f"dbname={settings.db_name} "
        f"user={settings.db_user} "
        f"password={settings.db_password}"
    )


@contextmanager
def get_conn():
    with psycopg.connect(_dsn()) as conn:
        yield conn


def _first_existing_table(conn: psycopg.Connection, candidates: list[str]) -> str | None:
    with conn.cursor() as cur:
        for name in candidates:
            cur.execute(
                """
                SELECT EXISTS (
                    SELECT 1
                    FROM information_schema.tables
                    WHERE table_schema = 'public' AND table_name = %s
                )
                """,
                (name,),
            )
            if cur.fetchone()[0]:
                return name
    return None


def get_client_profile(client_id: int) -> dict[str, Any] | None:
    sql = """
    SELECT
        c.id,
        c.name,
        c.resident_id,
        c.age,
        c.gender,
        c.education,
        c.desired_job,
        c.competency,
        c.address,
        c.university,
        c.major
    FROM clients c
    WHERE c.id = %s
    LIMIT 1
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (client_id,))
            row = cur.fetchone()
    if row is None:
        return None
    return {
        "clientId": int(row[0]),
        "name": row[1],
        "residentId": row[2],
        "age": row[3],
        "gender": row[4],
        "education": row[5],
        "desiredJob": row[6],
        "competency": row[7],
        "address": row[8],
        "university": row[9],
        "major": row[10],
    }


def get_all_client_profiles() -> list[dict[str, Any]]:
    sql = """
    SELECT
        c.id,
        c.name,
        c.resident_id,
        c.age,
        c.gender,
        c.education,
        c.desired_job,
        c.competency,
        c.address,
        c.university,
        c.major
    FROM clients c
    ORDER BY c.id
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
            rows = cur.fetchall()
    return [
        {
            "clientId": int(row[0]),
            "name": row[1],
            "residentId": row[2],
            "age": row[3],
            "gender": row[4],
            "education": row[5],
            "desiredJob": row[6],
            "competency": row[7],
            "address": row[8],
            "university": row[9],
            "major": row[10],
        }
        for row in rows
    ]


def update_client_embedding(client_id: int, embedding: list[float]) -> None:
    vector_text = "[" + ",".join(str(v) for v in embedding) + "]"
    sql = """
    UPDATE clients
    SET embedding = (%s)::vector
    WHERE id = %s
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (vector_text, client_id))
        conn.commit()


def get_client_embedding_source_hash(client_id: int) -> str | None:
    """embedding_source_hash 컬럼이 없으면 None 반환(에러 없음)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT embedding_source_hash
                    FROM clients
                    WHERE id = %s
                    """,
                    (client_id,),
                )
                row = cur.fetchone()
        return row[0] if row else None
    except Exception:
        return None


def set_client_embedding_source_hash(client_id: int, source_hash: str) -> None:
    """embedding_source_hash 컬럼이 없으면 무시(에러 없음)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE clients
                    SET embedding_source_hash = %s
                    WHERE id = %s
                    """,
                    (source_hash, client_id),
                )
            conn.commit()
    except Exception:
        pass


def compute_embedding_source_hash(embedding_text: str) -> str:
    return hashlib.sha256(embedding_text.encode("utf-8")).hexdigest()


def get_client_embedding_vector(client_id: int) -> list[float] | None:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT embedding
                FROM clients
                WHERE id = %s
                """,
                (client_id,),
            )
            row = cur.fetchone()
    if not row or row[0] is None:
        return None

    value = row[0]
    if isinstance(value, list):
        return [float(v) for v in value]
    if isinstance(value, tuple):
        return [float(v) for v in value]
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            text = text[1:-1]
        if not text:
            return None
        return [float(v) for v in text.split(",") if v.strip()]
    return None


def search_similar_cases(
    embedding: list[float],
    top_k: int = 5,
    exclude_client_id: int | None = None,
) -> list[dict[str, Any]]:
    vector_text = "[" + ",".join(str(v) for v in embedding) + "]"
    with get_conn() as conn:
        employment_table = _first_existing_table(conn, ["employment", "employments"])
        training_table = _first_existing_table(conn, ["training", "trainings"])
        consultation_table = _first_existing_table(conn, ["consultation", "consultations"])

        employment_select = (
            "e.job_title, e.company_name, e.salary"
            if employment_table
            else "NULL::varchar AS job_title, NULL::varchar AS company_name, NULL::integer AS salary"
        )
        employment_join = (
            f"""
            INNER JOIN LATERAL (
                SELECT e2.job_title, e2.company_name, e2.salary
                FROM {employment_table} e2
                WHERE e2.client_id = c.id AND e2.job_title IS NOT NULL AND e2.job_title <> ''
                ORDER BY e2.id DESC
                LIMIT 1
            ) e ON TRUE
            """
            if employment_table
            else ""
        )
        consultation_select = (
            "con.summary AS consultation_summary"
            if consultation_table
            else "NULL::text AS consultation_summary"
        )
        consultation_join = (
            f"""
            LEFT JOIN LATERAL (
                SELECT c2.summary
                FROM {consultation_table} c2
                WHERE c2.client_id = c.id
                ORDER BY c2.id DESC
                LIMIT 1
            ) con ON TRUE
            """
            if consultation_table
            else ""
        )
        trainings_select = (
            f"""
            COALESCE(
                (
                    SELECT json_agg(sub.course_name ORDER BY sub.id DESC)
                    FROM (
                        SELECT DISTINCT ON (t.course_name) t.course_name, t.id
                        FROM {training_table} t
                        WHERE t.client_id = ranked.client_id
                        ORDER BY t.course_name, t.id DESC
                    ) sub
                ),
                '[]'::json
            ) AS trainings
            """
            if training_table
            else "'[]'::json AS trainings"
        )

        sql = f"""
        WITH ranked AS (
            SELECT
                c.id AS client_id,
                c.desired_job,
                c.competency,
                c.education,
                c.major,
                c.university,
                c.age,
                c.gender,
                c.embedding <=> (%s)::vector AS distance,
                {consultation_select},
                {employment_select}
            FROM clients c
            {consultation_join}
            {employment_join}
            WHERE c.embedding IS NOT NULL
              AND (%s IS NULL OR c.id <> %s)
            ORDER BY c.embedding <=> (%s)::vector
            LIMIT %s
        )
        SELECT
            ranked.*,
            {trainings_select}
        FROM ranked;
        """
        with conn.cursor() as cur:
            cur.execute(
                sql,
                (
                    vector_text,
                    exclude_client_id,
                    exclude_client_id,
                    vector_text,
                    top_k,
                ),
            )
            rows = cur.fetchall()

    def _dedup_trainings(raw: Any) -> list:
        if isinstance(raw, str):
            raw = json.loads(raw)
        if not isinstance(raw, list):
            return []
        seen: set[str] = set()
        out: list = []
        for t in raw:
            s = str(t).strip() if t is not None else ""
            if s and s not in seen:
                seen.add(s)
                out.append(t)
        return out

    results: list[dict[str, Any]] = []
    for row in rows:
        trainings = _dedup_trainings(row[13])
        distance = float(row[8]) if row[8] is not None else 1.0
        score = max(0.0, 1.0 - distance)
        results.append(
            {
                "clientId": int(row[0]),
                "desiredJob": row[1],
                "competency": row[2],
                "education": row[3],
                "major": row[4],
                "university": row[5],
                "age": row[6] if isinstance(row[6], int) else None,
                "gender": row[7],
                "score": round(score, 6),
                "consultationSummary": row[9],
                "jobTitle": row[10],
                "companyName": row[11],
                "salary": row[12] if isinstance(row[12], int) else None,
                "trainings": trainings,
            }
        )
    return results


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    text = str(value).replace(",", "").strip()
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_bool_from_kor(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    if text in {"Y", "y", "예", "유", "O", "o", "1", "TRUE", "True", "true"}:
        return True
    if text in {"N", "n", "아니오", "무", "X", "x", "0", "FALSE", "False", "false"}:
        return False
    return None


def _allowance_from_kor(value: Any) -> str | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    if "지급" in text:
        return "PAID"
    if "미지급" in text:
        return "UNPAID"
    return "NONE"


def ingest_employment_training_from_excel(excel_path: str) -> dict[str, int]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    header_idx: dict[str, int] = {str(v).strip(): i for i, v in enumerate(headers) if v is not None}

    def col(name: str) -> int:
        if name not in header_idx:
            raise KeyError(f"엑셀 컬럼이 없습니다: {name}")
        return header_idx[name]

    idx_name = col("참여자\n이름")
    idx_resident = col("주민등록번호\n(숫자만 입력)")
    idx_course_name = col("직업훈련_훈련과정명")
    idx_train_start = col("직업훈련_훈련개강일")
    idx_train_end = col("직업훈련_훈련종료일")
    idx_allowance = col("직업훈련_훈련수당")
    idx_completed = col("훈련_수료\n여부")
    idx_consultation = col("상담 내역")
    idx_company = col("취업처_훈련수당")
    idx_job_title = col("취업직무_훈련수당")
    idx_salary = col("급여\n(전산망금액)_훈련수당")
    idx_employ_date = col("취업일자_훈련수당")
    idx_resign_date = col("퇴사일_훈련수당")

    clients_matched = 0
    employment_inserted = 0
    training_inserted = 0
    consultation_inserted = 0

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS consultation (
                    id BIGSERIAL PRIMARY KEY,
                    client_id BIGINT NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                    consult_date DATE,
                    raw_text TEXT,
                    summary TEXT,
                    iap_date DATE,
                    iap_period INTEGER,
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS training (
                    id BIGSERIAL PRIMARY KEY,
                    client_id BIGINT NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                    course_name VARCHAR(500),
                    start_date DATE,
                    end_date DATE,
                    allowance VARCHAR(100),
                    completed BOOLEAN,
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS employment (
                    id BIGSERIAL PRIMARY KEY,
                    client_id BIGINT NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                    company_name VARCHAR(255),
                    job_title VARCHAR(255),
                    salary INTEGER,
                    employ_date DATE,
                    resign_date DATE,
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )
                """
            )
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[idx_name]
                resident = row[idx_resident]
                if not name or not resident:
                    continue

                resident_digits = "".join(ch for ch in str(resident) if ch.isdigit())
                if len(resident_digits) != 13:
                    continue
                resident_id = resident_digits[:6] + "-" + resident_digits[6:]

                cur.execute(
                    """
                    SELECT id
                    FROM clients
                    WHERE name = %s AND resident_id = %s
                    LIMIT 1
                    """,
                    (str(name).strip(), resident_id),
                )
                found = cur.fetchone()
                if not found:
                    continue
                client_id = int(found[0])
                clients_matched += 1

                consultation_text = row[idx_consultation]
                if consultation_text:
                    cur.execute(
                        """
                        INSERT INTO consultation (
                            client_id, summary, created_at, updated_at
                        )
                        VALUES (%s, %s, NOW(), NOW())
                        """,
                        (
                            client_id,
                            str(consultation_text).strip(),
                        ),
                    )
                    consultation_inserted += 1

                course_name = row[idx_course_name]
                if course_name:
                    start_date = _parse_date(row[idx_train_start])
                    end_date = _parse_date(row[idx_train_end])
                    allowance = _allowance_from_kor(row[idx_allowance])
                    completed = _parse_bool_from_kor(row[idx_completed])
                    cur.execute(
                        """
                        INSERT INTO training (
                            client_id, course_name, start_date, end_date,
                            allowance, completed, created_at, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                        """,
                        (
                            client_id,
                            str(course_name).strip(),
                            start_date,
                            end_date,
                            allowance,
                            completed,
                        ),
                    )
                    training_inserted += 1

                company_name = row[idx_company]
                job_title = row[idx_job_title]
                salary = _parse_int(row[idx_salary])
                employ_date = _parse_date(row[idx_employ_date])
                resign_date = _parse_date(row[idx_resign_date])
                if company_name or job_title or salary or employ_date or resign_date:
                    cur.execute(
                        """
                        INSERT INTO employment (
                            client_id, company_name, job_title, salary,
                            employ_date, resign_date, created_at, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                        """,
                        (
                            client_id,
                            str(company_name).strip() if company_name else None,
                            str(job_title).strip() if job_title else None,
                            salary,
                            employ_date,
                            resign_date,
                        ),
                    )
                    employment_inserted += 1
        conn.commit()

    return {
        "clientsMatched": clients_matched,
        "employmentInserted": employment_inserted,
        "trainingInserted": training_inserted,
        "consultationInserted": consultation_inserted,
    }
